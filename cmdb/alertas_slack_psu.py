"""
Script de Alertas Slack para Gestão PSU GetNet
===============================================
Este script envia alertas automáticos para o canal Slack da ORAEX.

CONFIGURAÇÃO:
1. Crie um Incoming Webhook no Slack (veja guia_slack_webhook.md)
2. Cole a URL do webhook na variável SLACK_WEBHOOK_URL abaixo
3. Execute o script com: python alertas_slack_psu.py

AUTOMAÇÃO:
- No Windows, use o Agendador de Tarefas (Task Scheduler)
- Configure para executar às 17:00 de segunda a sexta
"""

import json
import urllib.request
from datetime import datetime, timedelta
import openpyxl
import os

# ============ CONFIGURAÇÃO ============
# A URL do webhook deve ser configurada como variável de ambiente
# Para uso local, crie um arquivo .env com: SLACK_WEBHOOK_URL=sua_url_aqui
# Para GitHub Actions, configure em Settings -> Secrets -> Actions
import os
SLACK_WEBHOOK_URL = os.environ.get("SLACK_WEBHOOK_URL", "")

# Caminho para a planilha de planejamento
import argparse
import sys

# Caminho para a planilha de planejamento relative to script
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANILHA_PATH = os.path.join(BASE_DIR, 'ORAEX_Planejamento_GetNet_2026.xlsx')

# ============ FUNÇÕES ============

def enviar_slack(mensagem: str, webhook_url: str = None) -> bool:
    """Envia uma mensagem para o Slack via webhook."""
    # Prioridade: Argumento -> Env Var -> Global
    url = webhook_url or os.environ.get("SLACK_WEBHOOK_URL", "")
    
    if not url:
        print("❌ ERRO: SLACK_WEBHOOK_URL não configurada.")
        return False
        
    try:
        payload = {
            "text": mensagem,
            "username": "PSU Bot GetNet",
            "icon_emoji": ":robot_face:"
        }
        data = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request(
            url, 
            data=data,
            headers={'Content-Type': 'application/json'}
        )
        urllib.request.urlopen(req)
        print(f"✅ Mensagem enviada para o Slack!")
        return True
    except Exception as e:
        print(f"❌ Erro ao enviar para Slack: {e}")
        return False


def carregar_servidores_criticos(planilha_path: str) -> list:
    """Carrega servidores com status 'Crítico' da planilha."""
    if not os.path.exists(planilha_path):
        print(f"❌ Planilha não encontrada: {planilha_path}")
        return []
        
    try:
        wb = openpyxl.load_workbook(planilha_path)
        ws = wb['Servidores']
        criticos = []
        # iter_rows começa da linha 1. Se cabeçalho está na 3, dados começam na 4.
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0]: continue # Pular vazio
            
            # Ajustar índices conforme colunas reais da planilha (A=0, E=4...)
            # Assumindo: Hostname(0), Ambiente(1), PSU(2), Status(4)
            status_val = row[4] if len(row) > 4 else None
            
            if status_val == 'Crítico':
                criticos.append({
                    'hostname': row[0],
                    'ambiente': row[1],
                    'psu_atual': row[2],
                    'ultima_atualizacao': row[5] if len(row) > 5 else 'N/A'
                })
        return criticos
    except Exception as e:
        print(f"Erro ao carregar planilha: {e}")
        return []


def lembrete_diario():
    """Envia lembrete diário às 17:00 com as atividades da noite."""
    # TODO: Ler atividades reais da planilha de 'GMUDs' para hoje
    hoje = datetime.now()
    data_formatada = hoje.strftime('%d/%m/%Y')
    
    mensagem = f"""
:calendar: *LEMBRETE DIÁRIO PSU - {data_formatada}*

:clock6: *Janela de Execução Hoje:*

:small_blue_diamond: *DEV* (18:00 - 03:00): Verificar Planilha
:small_blue_diamond: *HML* (18:00 - 03:00): Verificar Planilha
:small_blue_diamond: *PROD* (22:00 - 05:00): Verificar Planilha

:warning: *Lembretes:*
• Verificar aprovação do plantonista antes de iniciar
• Validar acesso aos servidores
• Atualizar status na planilha "ORAEX_Planejamento_GetNet_2026.xlsx".
• Preencher coluna "DESIGNADO A".

:rocket: Bom trabalho, equipe!
"""
    return enviar_slack(mensagem)


def alerta_servidores_criticos():
    """Envia alerta sobre servidores em estado crítico."""
    criticos = carregar_servidores_criticos(PLANILHA_PATH)
    
    if not criticos:
        print("Nenhum servidor crítico encontrado.")
        return True
    
    lista_servidores = "\n".join([
        f"• `{s['hostname']}` - {s['ambiente']} - PSU {s['psu_atual']}"
        for s in criticos[:15] # Limitar a 15 para não spammar
    ])
    
    truncado = "... (e mais)" if len(criticos) > 15 else ""
    
    mensagem = f"""
:rotating_light: *ALERTA: SERVIDORES CRÍTICOS*

Os seguintes servidores estão com PSU desatualizado (3+ quarters):

{lista_servidores}
{truncado}

:point_right: *Ação necessária:* Priorizar atualização destes hosts.
"""
    return enviar_slack(mensagem)


def resumo_semanal():
    """Envia resumo semanal às segundas-feiras."""
    hoje = datetime.now()
    semana_passada = hoje - timedelta(days=7)
    
    mensagem = f"""
:bar_chart: *RESUMO SEMANAL PSU - Semana {semana_passada.strftime('%d/%m')} a {hoje.strftime('%d/%m/%Y')}*

Acesse o Dashboard completo para ver os indicadores:
https://jonathanfferreira.github.io/oraex/

:memo: Não esqueça de atualizar a aba 'Execução' na planilha.
"""
    return enviar_slack(mensagem)


def testar_conexao():
    """Testa a conexão com o Slack."""
    mensagem = ":wave: *Teste de conexão do Bot PSU GetNet!*\n\nSe você está vendo esta mensagem, a integração está funcionando! :white_check_mark:"
    return enviar_slack(mensagem)


# ============ EXECUÇÃO ============
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Gestão de Alertas Slack PSU')
    parser.add_argument('--action', choices=['test', 'daily', 'critical', 'weekly'], help='Ação a ser executada')
    args = parser.parse_args()

    print("=" * 50)
    print("Sistema de Alertas PSU GetNet")
    print("=" * 50)
    
    if args.action:
        # Modo não-interativo (Automação)
        if args.action == 'test': testar_conexao()
        elif args.action == 'daily': lembrete_diario()
        elif args.action == 'critical': alerta_servidores_criticos()
        elif args.action == 'weekly': resumo_semanal()
    else:
        # Modo Interativo (Menu)
        print("\nEscolha uma opção:")
        print("1. Testar conexão com Slack")
        print("2. Enviar lembrete diário")
        print("3. Enviar alerta de servidores críticos")
        print("4. Enviar resumo semanal")
        print("5. Sair")
        
        opcao = input("\nOpção: ").strip()
        
        if opcao == "1": testar_conexao()
        elif opcao == "2": lembrete_diario()
        elif opcao == "3": alerta_servidores_criticos()
        elif opcao == "4": resumo_semanal()
        elif opcao == "5": print("Até mais!")
        else: print("Opção inválida.")
